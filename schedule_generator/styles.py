from openpyxl.styles import PatternFill

GRAY = PatternFill(
    start_color="708090", end_color="708090", fill_type="solid"
    )
RED = PatternFill(
    start_color="B22222", end_color="B22222", fill_type="solid"
    )
GREEN = PatternFill(
    start_color="008000", end_color="008000", fill_type="solid"
    )
YELLOW = PatternFill(
    start_color="FFD700", end_color="FFD700", fill_type="solid"
    )


def apply_style(writer, sheet_name):
    ws = writer.book[sheet_name]

    weekday_row = 2
    for col in range(2, ws.max_column + 1):
        cell = ws.cell(row=weekday_row,  column=col)
        if cell.value in ("Сб", "Вс"):
            cell.fill = RED
            ws.cell(row=2, column=col).fill = RED

    for row in range(3, ws.max_row + 1):
        for col in range(2, ws.max_column + 1):
            if ws.cell(row=row, column=col).value == "В":
                ws.cell(row=row, column=col).fill = RED
            elif ws.cell(row=row, column=col).value == "Д2":
                ws.cell(row=row, column=col).fill = YELLOW
            elif ws.cell(row=row, column=col).value == "Д1":
                ws.cell(row=row, column=col).fill = GREEN
            elif ws.cell(row=row, column=col).value == "":
                ws.cell(row=row, column=col).fill = GRAY
